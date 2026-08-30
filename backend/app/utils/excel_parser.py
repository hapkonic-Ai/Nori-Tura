"""Excel/CSV patient import parser."""

import io
from typing import List, Dict, Any, Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


REQUIRED_COLUMNS = {"name", "age", "gender", "parent_name", "parent_phone"}
OPTIONAL_COLUMNS = {
    "blood_group",
    "allergies",
    "doctor_phone",
    "doctor_id",
    "hospital_name",
    "hospital_id",
}
ALLOWED_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize raw row values: strip strings, convert blanks to None."""
    normalized: Dict[str, Any] = {}
    for key, value in row.items():
        if value is None:
            normalized[key] = None
            continue
        if isinstance(value, str):
            stripped = value.strip()
            normalized[key] = stripped if stripped else None
        else:
            normalized[key] = value
    return normalized


def _coerce_int(value: Any, field: str) -> int:
    """Coerce a value to int with a clear error."""
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"'{field}' must be an integer, got {value!r}") from exc


def parse_excel_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse an XLSX or CSV file into a list of normalized row dicts."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("Excel file has no active worksheet")

    # First row as headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value.strip().lower() if isinstance(cell.value, str) else cell.value)

    if not headers or all(h is None for h in headers):
        raise ValueError("No headers found in Excel file")

    rows: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_data = {}
        for header, value in zip(headers, row):
            if header is not None:
                row_data[header] = value
        rows.append(_normalize_row(row_data))

    return rows


def validate_and_prepare_rows(
    rows: List[Dict[str, Any]],
    default_doctor_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Validate rows and return a list of prepared patient dicts.
    Raises ValueError with a detailed message on first failure (strict mode).
    """
    if not rows:
        raise ValueError("No data rows found in the uploaded file")

    errors: List[str] = []
    prepared: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=1):
        row_errors: List[str] = []

        # Unknown columns
        unknown = set(row.keys()) - ALLOWED_COLUMNS
        if unknown:
            row_errors.append(f"unknown columns: {', '.join(sorted(unknown))}")

        # Required columns
        missing = REQUIRED_COLUMNS - set(row.keys())
        for col in missing:
            if row.get(col) is None:
                row_errors.append(f"missing required column '{col}'")

        if row_errors:
            errors.append(f"Row {idx}: {'; '.join(row_errors)}")
            continue

        # Value validation
        name = row.get("name")
        if not name:
            row_errors.append("name cannot be empty")

        try:
            age = _coerce_int(row.get("age"), "age")
            if age < 0 or age > 150:
                row_errors.append("age must be between 0 and 150")
        except ValueError as exc:
            row_errors.append(str(exc))

        gender = (row.get("gender") or "").strip()
        if not gender:
            row_errors.append("gender cannot be empty")

        parent_name = row.get("parent_name")
        if not parent_name:
            row_errors.append("parent_name cannot be empty")

        parent_phone = row.get("parent_phone")
        if not parent_phone:
            row_errors.append("parent_phone cannot be empty")
        elif not isinstance(parent_phone, str) or not parent_phone.startswith("+91") or len(parent_phone) != 13:
            row_errors.append("parent_phone must be +91 followed by 10 digits")

        doctor_id = row.get("doctor_id") or default_doctor_id
        doctor_phone = row.get("doctor_phone")
        if not doctor_id and not doctor_phone:
            row_errors.append("either doctor_id or doctor_phone is required")

        if row_errors:
            errors.append(f"Row {idx}: {'; '.join(row_errors)}")
            continue

        prepared.append({
            "name": name,
            "age": age,
            "gender": gender,
            "blood_group": row.get("blood_group"),
            "allergies": row.get("allergies"),
            "parent_name": parent_name,
            "parent_phone": parent_phone,
            "doctor_id": doctor_id,
            "doctor_phone": doctor_phone,
            "hospital_id": row.get("hospital_id"),
            "hospital_name": row.get("hospital_name"),
        })

    if errors:
        raise ValueError("\n".join(errors))

    return prepared
